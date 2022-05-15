##### Data Driven Testing using Excel #####

import openpyxl   # for excel operations

def excel_operations():

    excel_path = 'C:\\Users\\ranje\\Downloads\\emp_data.xlsx'   # excel file location
    workbook = openpyxl.load_workbook(excel_path)                          # loading the excel file

    # activeSheet = workbook.active       # loads active sheet
    sheet = workbook.get_sheet_by_name("Sheet2")     # loading specific sheet

    ## finding number of rows and columns
    # rows = activeSheet.max_row
    # print(rows)
    #
    # cols = activeSheet.max_column
    # print(cols)
    #
    # ## 1. Reading Data from Excel file    #### for reading data we always need to find-out number of rows and columns
    #
    # for r in range(1,rows+1):
    #     for c in range(1,cols+1):
    #         print(activeSheet.cell(row=r,column=c).value,end ="          ")
    #     print()


    ### 2. Writing data into excel file

    for r in range(1,6):
        for c in range(1,4):
            sheet.cell(row = r,column = c).value = "welcome"

    workbook.save(excel_path)    ### To save the inserted data



## 3. Data Driven Test cases




    return workbook

excel_operations()

