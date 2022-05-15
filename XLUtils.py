import openpyxl

def getRowCount(file_path,sheetName):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)

def getColumnCount(file_path,sheetName):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)

def readData(file_path,sheetName,rowNum,colnumNo):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rowNum,column=colnumNo).value

def writeData(file_path,sheetName,rowNum,columnNo,data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowNum,column=columnNo).value = data
    workbook.save(file_path)

