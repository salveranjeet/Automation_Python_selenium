import openpyxl

file = "E:\python_code\Automation_Python_selenium\TestData\data.xlsx"

#loading workbook from file
workbook = openpyxl.load_workbook(file)

#Extracting sheet from workbook
sheet = workbook["Sheet1"]

#Number of rows,cols
rows = sheet.max_row
cols = sheet.max_column

#Reading all the rows and columns data from the excel sheet

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end='    ')
    print()


