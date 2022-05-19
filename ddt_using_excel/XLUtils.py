import openpyxl

# for applying colours on the cells
from openpyxl.styles import PatternFill

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)

def readData(file,sheetName,rowNum,colnumNo):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNum,column=colnumNo).value

def writeData(file,sheetName,rowNum,columnNo,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook["Sheet1"]
    sheet.cell(row=rowNum,column=columnNo).value = data
    workbook.save(file)

def fillGreenColor(file,sheetName,rowNum,columnNo):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212', end_color='60b212',fill_type='solid')
    sheet.cell(rowNum,columnNo).fill = greenFill
    workbook.save(file)

def fillRedColor(file,sheetName,rowNum,columnNo):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='ff0000', end_color='ff0000',fill_type='solid')
    sheet.cell(rowNum,columnNo).fill = redFill
    workbook.save(file)
