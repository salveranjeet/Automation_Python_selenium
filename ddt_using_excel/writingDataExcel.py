import openpyxl

# file = "E:\python_code\Automation_Python_selenium\TestData\data1.xlsx"

file = "E:\python_code\Automation_Python_selenium\TestData\login1.xlsx"

#loading workbook from file
workbook = openpyxl.load_workbook(file)

#Extracting sheet from workbook
sheet = workbook.active    #will give active sheet in excel(when one sheet is in excel)

#Number of rows,cols
rows = sheet.max_row
cols = sheet.max_column

#Writing data in Excel (same value will be added everywhere)
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value= "DDT Excel"
#
# workbook.save(file)

# Writing diff data in cells
sheet.cell(2,4).value="As Expected"
sheet.cell(3,4).value="Not As Expected"
sheet.cell(4,4).value="Not As Expected"
sheet.cell(5,4).value="As Expected"

workbook.save(file)
